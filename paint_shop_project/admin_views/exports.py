"""
Экспорт отчетов в Excel
"""
from datetime import timedelta
from decimal import Decimal
from io import BytesIO

from django.contrib.auth.decorators import user_passes_test
from django.db.models import Count, DecimalField, ExpressionWrapper, F, Q, Sum
from django.db.models.functions import TruncDay
from django.http import HttpResponse
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.utils.translation import gettext_lazy as _
from django.views.generic import TemplateView

from ..models import Order, OrderItem, Product, User, ProductBatch, BatchAuditLog, Category


def is_staff(user):
    return user.is_staff


def _get_excel_styles():
    """Возвращает набор стилей для Excel файлов"""
    try:
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        return {
            'header_fill': PatternFill(start_color='1a2a6c', end_color='1a2a6c', fill_type='solid'),
            'header_font': Font(bold=True, color='FFFFFF', size=12),
            'header_alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'border': Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            ),
            'alt_fill': PatternFill(start_color='f0f4ff', end_color='f0f4ff', fill_type='solid'),
            'total_fill': PatternFill(start_color='e8f0fe', end_color='e8f0fe', fill_type='solid'),
            'total_font': Font(bold=True, size=11),
            'data_font': Font(size=10),
            'get_column_letter': get_column_letter,
        }
    except ImportError:
        return None


def _apply_excel_header_style(ws, headers, styles):
    """Применяет стили к заголовкам Excel таблицы"""
    if not styles:
        return
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.alignment = styles['header_alignment']
        cell.border = styles['border']
        # Автоширина колонок
        ws.column_dimensions[styles['get_column_letter'](col)].width = max(len(str(header)) + 2, 15)
    
    ws.row_dimensions[1].height = 25


def _apply_excel_row_style(ws, row, num_cols, styles, is_alt=False, is_total=False):
    """Применяет стили к строке данных Excel"""
    if not styles:
        return
    
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = styles['border']
        
        if is_total:
            cell.font = styles['total_font']
            cell.fill = styles['total_fill']
        else:
            cell.font = styles['data_font']
            if is_alt:
                cell.fill = styles['alt_fill']


@method_decorator(user_passes_test(is_staff), name='dispatch')
class ExportReportsView(TemplateView):
    """Экспорт отчетов в Excel"""
    template_name = 'admin/export_reports.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = timezone.now().date()
        context.update({
            'title': _('Экспорт отчетов'),
            'today': today,
            'last_month': today - timedelta(days=30),
            'next_month': today + timedelta(days=30),
        })
        return context

    def post(self, request, *args, **kwargs):
        """Генерация и скачивание отчета"""
        report_type = request.POST.get('report_type')
        
        if report_type == 'sales':
            return self._export_sales_report(request)
        elif report_type == 'orders':
            return self._export_orders_report(request)
        elif report_type == 'products':
            return self._export_products_report(request)
        elif report_type == 'customers':
            return self._export_customers_report(request)
        elif report_type == 'batches_expiring':
            return self._export_batches_expiring_report(request)
        elif report_type == 'batches_history':
            return self._export_batches_history_report(request)
        elif report_type == 'batches_losses':
            return self._export_batches_losses_report(request)
        elif report_type == 'batches_by_category':
            return self._export_batches_by_category_report(request)
        elif report_type == 'dashboard_pdf':
            return self._export_dashboard_pdf(request)
        elif report_type == 'all_metrics_export':
            return self._export_all_metrics(request)
        elif report_type == 'all_metrics_import':
            return self._import_all_metrics(request)
        elif report_type == 'export_styles':
            return self._export_styles_css(request)
        else:
            return self.get(request)

    def _export_sales_report(self, request):
        """Экспорт отчета по продажам"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Продажи"
            
            # Стили
            header_fill = PatternFill(start_color='1a2a6c', end_color='1a2a6c', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=12)
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            border_style = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            
            alt_fill = PatternFill(start_color='f0f4ff', end_color='f0f4ff', fill_type='solid')
            total_fill = PatternFill(start_color='e8f0fe', end_color='e8f0fe', fill_type='solid')
            total_font = Font(bold=True, size=11)
            
            # Заголовки
            headers = ['Дата', 'Заказов', 'Выручка', 'Средний чек']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border_style
                # Автоширина колонок
                ws.column_dimensions[get_column_letter(col)].width = max(len(header) + 2, 15)
            
            # Данные
            now = timezone.now()
            days = int(request.POST.get('days', 30))
            start_date = now - timedelta(days=days)
            
            daily_data = (
                Order.objects.filter(
                    order_date__gte=start_date,
                    status='delivered'
                )
                .annotate(day=TruncDay('order_date'))
                .values('day')
                .annotate(
                    orders=Count('id'),
                    revenue=Sum('total_amount')
                )
                .order_by('day')
            )
            
            row = 2
            for idx, item in enumerate(daily_data):
                revenue = float(item['revenue'] or 0)
                orders = item['orders']
                avg = revenue / orders if orders > 0 else 0
                
                # Чередующиеся цвета строк
                row_fill = alt_fill if idx % 2 == 0 else None
                
                ws.cell(row=row, column=1, value=item['day'].strftime('%d.%m.%Y'))
                ws.cell(row=row, column=2, value=orders)
                ws.cell(row=row, column=3, value=revenue).number_format = '#,##0.00 ₽'
                ws.cell(row=row, column=4, value=avg).number_format = '#,##0.00 ₽'
                
                # Применяем стили
                for col in range(1, 5):
                    cell = ws.cell(row=row, column=col)
                    cell.border = border_style
                    if row_fill:
                        cell.fill = row_fill
                    if col == 1:
                        cell.alignment = Alignment(horizontal='left')
                    else:
                        cell.alignment = Alignment(horizontal='right')
                
                row += 1
            
            # Итоговая строка
            total_row = row
            ws.cell(row=total_row, column=1, value='ИТОГО').font = total_font
            ws.cell(row=total_row, column=1).fill = total_fill
            ws.cell(row=total_row, column=1).border = border_style
            ws.cell(row=total_row, column=1).alignment = Alignment(horizontal='left')
            
            total_orders = sum(item['orders'] for item in daily_data)
            total_revenue = sum(float(item['revenue'] or 0) for item in daily_data)
            total_avg = total_revenue / total_orders if total_orders > 0 else 0
            
            ws.cell(row=total_row, column=2, value=total_orders).font = total_font
            ws.cell(row=total_row, column=2).fill = total_fill
            ws.cell(row=total_row, column=2).border = border_style
            ws.cell(row=total_row, column=2).alignment = Alignment(horizontal='right')
            
            ws.cell(row=total_row, column=3, value=total_revenue).font = total_font
            ws.cell(row=total_row, column=3).number_format = '#,##0.00 ₽'
            ws.cell(row=total_row, column=3).fill = total_fill
            ws.cell(row=total_row, column=3).border = border_style
            ws.cell(row=total_row, column=3).alignment = Alignment(horizontal='right')
            
            ws.cell(row=total_row, column=4, value=total_avg).font = total_font
            ws.cell(row=total_row, column=4).number_format = '#,##0.00 ₽'
            ws.cell(row=total_row, column=4).fill = total_fill
            ws.cell(row=total_row, column=4).border = border_style
            ws.cell(row=total_row, column=4).alignment = Alignment(horizontal='right')
            
            # Высота строки заголовка
            ws.row_dimensions[1].height = 25
            
            # Сохранение
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename=sales_report_{timezone.now().strftime("%Y%m%d")}.xlsx'
            return response
            
        except ImportError:
            from django.contrib import messages
            messages.error(request, _('Библиотека openpyxl не установлена. Установите: pip install openpyxl'))
            return self.get(request)

    def _export_orders_report(self, request):
        """Экспорт отчета по заказам"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Заказы"
            
            styles = _get_excel_styles()
            headers = ['ID', 'Дата', 'Клиент', 'Статус', 'Сумма', 'Товаров']
            _apply_excel_header_style(ws, headers, styles)
            
            days = int(request.POST.get('days', 30))
            start_date = timezone.now() - timedelta(days=days)
            
            orders = Order.objects.filter(order_date__gte=start_date).select_related('user')
            
            row = 2
            for idx, order in enumerate(orders):
                items_count = order.orderitem_set.count()
                ws.cell(row=row, column=1, value=order.id)
                ws.cell(row=row, column=2, value=order.order_date.strftime('%d.%m.%Y %H:%M'))
                ws.cell(row=row, column=3, value=order.user.username if order.user else '—')
                ws.cell(row=row, column=4, value=order.get_status_display())
                ws.cell(row=row, column=5, value=float(order.total_amount)).number_format = '#,##0.00 ₽'
                ws.cell(row=row, column=6, value=items_count)
                
                _apply_excel_row_style(ws, row, len(headers), styles, is_alt=(idx % 2 == 0))
                
                # Выравнивание
                ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
                ws.cell(row=row, column=2).alignment = Alignment(horizontal='left')
                ws.cell(row=row, column=3).alignment = Alignment(horizontal='left')
                ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')
                ws.cell(row=row, column=5).alignment = Alignment(horizontal='right')
                ws.cell(row=row, column=6).alignment = Alignment(horizontal='center')
                
                row += 1
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename=orders_report_{timezone.now().strftime("%Y%m%d")}.xlsx'
            return response
            
        except ImportError:
            from django.contrib import messages
            messages.error(request, _('Библиотека openpyxl не установлена.'))
            return self.get(request)

    def _export_products_report(self, request):
        """Экспорт отчета по товарам"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Товары"
            
            styles = _get_excel_styles()
            headers = ['ID', 'Название', 'Категория', 'Цена', 'Продано шт.', 'Выручка', 'Заказов']
            _apply_excel_header_style(ws, headers, styles)
            
            days = int(request.POST.get('days', 30))
            start_date = timezone.now() - timedelta(days=days)
            
            # Агрегируем данные по товарам
            products_data = list((
                OrderItem.objects.filter(
                    order__order_date__gte=start_date,
                    order__status='delivered'
                )
                .values('product__id', 'product__name', 'product__category__name', 'product__price')
                .annotate(
                    total_quantity=Sum('quantity'),
                    total_revenue=Sum(ExpressionWrapper(F('quantity') * F('price_per_unit'), output_field=DecimalField())),
                    orders_count=Count('order', distinct=True)
                )
                .order_by('-total_revenue')
            ))
            
            row = 2
            for idx, item in enumerate(products_data):
                ws.cell(row=row, column=1, value=item['product__id'] or '—')
                ws.cell(row=row, column=2, value=item['product__name'] or '—')
                ws.cell(row=row, column=3, value=item['product__category__name'] or 'Без категории')
                ws.cell(row=row, column=4, value=float(item['product__price'] or 0)).number_format = '#,##0.00 ₽'
                ws.cell(row=row, column=5, value=item['total_quantity'] or 0)
                ws.cell(row=row, column=6, value=float(item['total_revenue'] or 0)).number_format = '#,##0.00 ₽'
                ws.cell(row=row, column=7, value=item['orders_count'] or 0)
                
                _apply_excel_row_style(ws, row, len(headers), styles, is_alt=(idx % 2 == 0))
                
                # Выравнивание
                ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
                ws.cell(row=row, column=2).alignment = Alignment(horizontal='left')
                ws.cell(row=row, column=3).alignment = Alignment(horizontal='left')
                ws.cell(row=row, column=4).alignment = Alignment(horizontal='right')
                ws.cell(row=row, column=5).alignment = Alignment(horizontal='right')
                ws.cell(row=row, column=6).alignment = Alignment(horizontal='right')
                ws.cell(row=row, column=7).alignment = Alignment(horizontal='center')
                
                row += 1
            
            # Итоговая строка
            if row > 2:
                total_row = row
                ws.cell(row=total_row, column=1, value='ИТОГО')
                ws.cell(row=total_row, column=5, value=sum(item['total_quantity'] or 0 for item in products_data))
                ws.cell(row=total_row, column=6, value=sum(float(item['total_revenue'] or 0) for item in products_data))
                ws.cell(row=total_row, column=7, value=sum(item['orders_count'] or 0 for item in products_data))
                
                _apply_excel_row_style(ws, total_row, len(headers), styles, is_total=True)
                ws.cell(row=total_row, column=1).alignment = Alignment(horizontal='left')
                ws.cell(row=total_row, column=5).alignment = Alignment(horizontal='right')
                ws.cell(row=total_row, column=6).alignment = Alignment(horizontal='right')
                ws.cell(row=total_row, column=6).number_format = '#,##0.00 ₽'
                ws.cell(row=total_row, column=7).alignment = Alignment(horizontal='center')
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename=products_report_{timezone.now().strftime("%Y%m%d")}.xlsx'
            return response
            
        except ImportError:
            from django.contrib import messages
            messages.error(request, _('Библиотека openpyxl не установлена.'))
            return self.get(request)

    def _export_customers_report(self, request):
        """Экспорт отчета по клиентам"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Клиенты"
            
            styles = _get_excel_styles()
            headers = ['ID', 'Имя пользователя', 'Email', 'Имя', 'Фамилия', 'Заказов', 'Выручка', 'Средний чек', 'Дата регистрации']
            _apply_excel_header_style(ws, headers, styles)
            
            days = int(request.POST.get('days', 30))
            start_date = timezone.now() - timedelta(days=days)
            
            # Агрегируем данные по клиентам
            customers_data = list((
                User.objects.filter(
                    is_staff=False,
                    orders__order_date__gte=start_date,
                    orders__status='delivered'
                )
                .annotate(
                    orders_count=Count('orders', distinct=True),
                    total_revenue=Sum('orders__total_amount'),
                    avg_order_value=ExpressionWrapper(
                        Sum('orders__total_amount') / Count('orders', distinct=True),
                        output_field=DecimalField()
                    )
                )
                .filter(orders_count__gt=0)
                .order_by('-total_revenue')
            ))
            
            row = 2
            for idx, customer in enumerate(customers_data):
                ws.cell(row=row, column=1, value=customer.id)
                ws.cell(row=row, column=2, value=customer.username)
                ws.cell(row=row, column=3, value=customer.email or '—')
                ws.cell(row=row, column=4, value=customer.first_name or '—')
                ws.cell(row=row, column=5, value=customer.last_name or '—')
                ws.cell(row=row, column=6, value=customer.orders_count or 0)
                ws.cell(row=row, column=7, value=float(customer.total_revenue or 0)).number_format = '#,##0.00 ₽'
                ws.cell(row=row, column=8, value=float(customer.avg_order_value or 0)).number_format = '#,##0.00 ₽'
                ws.cell(row=row, column=9, value=customer.date_joined.strftime('%d.%m.%Y') if customer.date_joined else '—')
                
                _apply_excel_row_style(ws, row, len(headers), styles, is_alt=(idx % 2 == 0))
                
                # Выравнивание
                ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
                ws.cell(row=row, column=2).alignment = Alignment(horizontal='left')
                ws.cell(row=row, column=3).alignment = Alignment(horizontal='left')
                ws.cell(row=row, column=4).alignment = Alignment(horizontal='left')
                ws.cell(row=row, column=5).alignment = Alignment(horizontal='left')
                ws.cell(row=row, column=6).alignment = Alignment(horizontal='center')
                ws.cell(row=row, column=7).alignment = Alignment(horizontal='right')
                ws.cell(row=row, column=8).alignment = Alignment(horizontal='right')
                ws.cell(row=row, column=9).alignment = Alignment(horizontal='left')
                
                row += 1
            
            # Итоговая строка
            if row > 2:
                total_row = row
                ws.cell(row=total_row, column=1, value='ИТОГО')
                ws.cell(row=total_row, column=6, value=sum(c.orders_count or 0 for c in customers_data))
                ws.cell(row=total_row, column=7, value=sum(float(c.total_revenue or 0) for c in customers_data))
                
                _apply_excel_row_style(ws, total_row, len(headers), styles, is_total=True)
                ws.cell(row=total_row, column=1).alignment = Alignment(horizontal='left')
                ws.cell(row=total_row, column=6).alignment = Alignment(horizontal='center')
                ws.cell(row=total_row, column=7).alignment = Alignment(horizontal='right')
                ws.cell(row=total_row, column=7).number_format = '#,##0.00 ₽'
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename=customers_report_{timezone.now().strftime("%Y%m%d")}.xlsx'
            return response
            
        except ImportError:
            from django.contrib import messages
            messages.error(request, _('Библиотека openpyxl не установлена.'))
            return self.get(request)

    def _export_dashboard_pdf(self, request):
        """Экспорт дашборда в PDF с графиками"""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            import matplotlib
            matplotlib.use('Agg')
            import matplotlib.pyplot as plt
            import matplotlib.font_manager as fm
            from datetime import datetime
            import os
            
            # Настройка шрифтов для кириллицы
            try:
                # Пытаемся использовать системные шрифты Windows
                font_paths = [
                    'C:/Windows/Fonts/arial.ttf',
                    'C:/Windows/Fonts/arialbd.ttf',
                    'C:/Windows/Fonts/times.ttf',
                    'C:/Windows/Fonts/timesbd.ttf',
                ]
                
                # Регистрируем шрифты если они доступны
                for font_path in font_paths:
                    if os.path.exists(font_path):
                        font_name = os.path.basename(font_path).replace('.ttf', '').replace('arial', 'Arial').replace('times', 'Times')
                        if 'bd' in font_path:
                            font_name += '-Bold'
                        try:
                            pdfmetrics.registerFont(TTFont(font_name, font_path))
                        except:
                            pass
            except:
                pass
            
            # Настройка matplotlib для кириллицы
            # Пытаемся найти шрифт с поддержкой кириллицы
            import matplotlib.font_manager as fm
            font_list = fm.findSystemFonts(fontpaths=None, fontext='ttf')
            cyrillic_fonts = ['DejaVu', 'Arial', 'Liberation', 'Calibri', 'Tahoma', 'Verdana']
            found_font = None
            
            for font_path in font_list:
                try:
                    font_name = fm.get_font(font_path).family_name
                    if any(cyr in font_name for cyr in cyrillic_fonts):
                        found_font = font_name
                        break
                except:
                    continue
            
            if found_font:
                plt.rcParams['font.family'] = found_font
            else:
                # Fallback на DejaVu Sans, который обычно есть
                plt.rcParams['font.family'] = ['DejaVu Sans', 'Arial', 'Liberation Sans', 'sans-serif']
            
            plt.rcParams['axes.unicode_minus'] = False
            plt.rcParams['font.size'] = 10
            
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
            elements = []
            styles = getSampleStyleSheet()
            
            # Создаем стили с поддержкой кириллицы
            try:
                title_font = 'Arial-Bold' if 'Arial-Bold' in pdfmetrics.getRegisteredFontNames() else 'Helvetica-Bold'
                normal_font = 'Arial' if 'Arial' in pdfmetrics.getRegisteredFontNames() else 'Helvetica'
            except:
                title_font = 'Helvetica-Bold'
                normal_font = 'Helvetica'
            
            # Заголовок
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#1a2a6c'),
                spaceAfter=30,
                alignment=TA_CENTER,
                fontName=title_font
            )
            
            # Стиль для описаний
            desc_style = ParagraphStyle(
                'Description',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.HexColor('#666666'),
                spaceAfter=12,
                alignment=TA_JUSTIFY,
                fontName=normal_font
            )
            
            # Стиль для подзаголовков
            subtitle_style = ParagraphStyle(
                'Subtitle',
                parent=styles['Heading2'],
                fontSize=16,
                textColor=colors.HexColor('#1a2a6c'),
                spaceAfter=12,
                spaceBefore=12,
                alignment=TA_LEFT,
                fontName=title_font
            )
            
            # Заголовок документа
            elements.append(Paragraph('Дашборд аналитики продаж', title_style))
            elements.append(Spacer(1, 0.1*inch))
            
            # Информация о периоде
            period = int(request.POST.get('period', 30))
            period_text = f'<b>Период анализа:</b> {period} дней'
            elements.append(Paragraph(period_text, desc_style))
            elements.append(Paragraph(f'<b>Дата формирования отчета:</b> {timezone.now().strftime("%d.%m.%Y %H:%M")}', desc_style))
            elements.append(Spacer(1, 0.2*inch))
            
            # Получаем данные
            period = int(request.POST.get('period', 30))
            now = timezone.now()
            start_date = now - timedelta(days=period)
            
            # Данные для графиков
            daily_data = (
                Order.objects.filter(
                    order_date__gte=start_date,
                    status='delivered'
                )
                .annotate(day=TruncDay('order_date'))
                .values('day')
                .annotate(
                    revenue=Sum('total_amount'),
                    orders=Count('id')
                )
                .order_by('day')
            )
            
            # Статистика по статусам
            orders_by_status = (
                Order.objects.filter(order_date__gte=start_date)
                .values('status')
                .annotate(count=Count('id'))
            )
            
            # Топ товары
            revenue_expr = ExpressionWrapper(
                F('quantity') * F('price_per_unit'),
                output_field=DecimalField(max_digits=16, decimal_places=2)
            )
            top_products = (
                OrderItem.objects.filter(
                    order__order_date__gte=start_date,
                    order__status='delivered'
                )
                .values('product__name')
                .annotate(
                    revenue=Sum(revenue_expr),
                    quantity=Sum('quantity'),
                    orders=Count('order', distinct=True)
                )
                .order_by('-revenue')[:10]
            )
            
            # График 1: Выручка по дням
            dates = [item['day'] for item in daily_data]
            revenues = [float(item['revenue'] or 0) for item in daily_data]
            
            if dates:
                elements.append(Paragraph('1. Выручка по дням', subtitle_style))
                elements.append(Paragraph(
                    'График показывает динамику ежедневной выручки за выбранный период. '
                    'Позволяет выявить тренды роста или падения продаж, а также определить наиболее прибыльные дни.',
                    desc_style
                ))
                
                fig, ax = plt.subplots(figsize=(10, 4.5))
                ax.plot(dates, revenues, marker='o', linewidth=2.5, color='#1a2a6c', markersize=6)
                ax.fill_between(dates, revenues, alpha=0.3, color='#1a2a6c')
                ax.set_title('Выручка по дням', fontsize=16, fontweight='bold', pad=15)
                ax.set_xlabel('Дата', fontsize=12, fontweight='bold')
                ax.set_ylabel('Выручка (₽)', fontsize=12, fontweight='bold')
                ax.grid(True, alpha=0.3, linestyle='--')
                plt.xticks(rotation=45, ha='right')
                
                # Добавляем аннотацию с максимальным значением
                if revenues:
                    max_idx = revenues.index(max(revenues))
                    max_revenue = max(revenues)
                    ax.annotate(f'Макс: {max_revenue:.2f} ₽', 
                              xy=(dates[max_idx], max_revenue),
                              xytext=(10, 10), textcoords='offset points',
                              bbox=dict(boxstyle='round,pad=0.5', facecolor='yellow', alpha=0.7),
                              arrowprops=dict(arrowstyle='->', connectionstyle='arc3,rad=0'))
                
                plt.tight_layout()
                
                chart1 = BytesIO()
                plt.savefig(chart1, format='png', dpi=110, bbox_inches='tight')
                plt.close()
                chart1.seek(0)
                
                from reportlab.platypus import Image as RLImage
                elements.append(RLImage(chart1, width=9*inch, height=4.3*inch))
                elements.append(Spacer(1, 0.3*inch))
            
            # График 2: Количество заказов
            orders_count = [item['orders'] for item in daily_data]
            if dates:
                elements.append(PageBreak())
                elements.append(Paragraph('2. Количество заказов по дням', subtitle_style))
                elements.append(Paragraph(
                    'Столбчатая диаграмма отображает количество заказов, оформленных каждый день. '
                    'Помогает анализировать активность покупателей и планировать загрузку службы доставки.',
                    desc_style
                ))
                
                fig, ax = plt.subplots(figsize=(10, 4.5))
                bars = ax.bar(dates, orders_count, color='#4f63d8', alpha=0.8, edgecolor='#1a2a6c', linewidth=1.5)
                ax.set_title('Количество заказов по дням', fontsize=16, fontweight='bold', pad=15)
                ax.set_xlabel('Дата', fontsize=12, fontweight='bold')
                ax.set_ylabel('Количество заказов', fontsize=12, fontweight='bold')
                ax.grid(True, alpha=0.3, axis='y', linestyle='--')
                plt.xticks(rotation=45, ha='right')
                
                # Добавляем значения на столбцы
                for bar in bars:
                    height = bar.get_height()
                    ax.text(bar.get_x() + bar.get_width()/2., height,
                           f'{int(height)}',
                           ha='center', va='bottom', fontweight='bold')
                
                plt.tight_layout()
                
                chart2 = BytesIO()
                plt.savefig(chart2, format='png', dpi=110, bbox_inches='tight')
                plt.close()
                chart2.seek(0)
                
                elements.append(RLImage(chart2, width=9*inch, height=4.3*inch))
                elements.append(Spacer(1, 0.3*inch))
            
            # График 3: Статусы заказов
            status_labels = []
            status_counts = []
            status_translation = {
                'created': 'Создан',
                'confirmed': 'Подтверждён',
                'ready': 'Готов',
                'in_transit': 'В пути',
                'delivered': 'Доставлен',
                'cancelled': 'Отменён'
            }
            
            for item in orders_by_status:
                status_labels.append(status_translation.get(item['status'], item['status']))
                status_counts.append(item['count'])
            
            if status_labels:
                elements.append(PageBreak())
                elements.append(Paragraph('3. Распределение заказов по статусам', subtitle_style))
                elements.append(Paragraph(
                    'Круговая диаграмма показывает распределение всех заказов по статусам выполнения. '
                    'Позволяет оценить эффективность обработки заказов и выявить узкие места в процессе доставки.',
                    desc_style
                ))
                
                fig, ax = plt.subplots(figsize=(8, 6))
                colors_list = ['#10b981', '#3b82f6', '#f59e0b', '#ef4444', '#8b5cf6', '#ec4899']
                wedges, texts, autotexts = ax.pie(
                    status_counts, 
                    labels=status_labels, 
                    autopct='%1.1f%%',
                    colors=colors_list[:len(status_labels)],
                    startangle=90,
                    textprops={'fontsize': 11, 'fontweight': 'bold'}
                )
                ax.set_title('Распределение заказов по статусам', fontsize=16, fontweight='bold', pad=20)
                
                # Улучшаем отображение процентов
                for autotext in autotexts:
                    autotext.set_color('white')
                    autotext.set_fontweight('bold')
                    autotext.set_fontsize(10)
                
                plt.tight_layout()
                
                chart3 = BytesIO()
                plt.savefig(chart3, format='png', dpi=110, bbox_inches='tight')
                plt.close()
                chart3.seek(0)
                
                elements.append(RLImage(chart3, width=7*inch, height=5*inch))
                elements.append(Spacer(1, 0.3*inch))
            
            # График 4: Топ-10 товаров
            if top_products:
                elements.append(PageBreak())
                elements.append(Paragraph('4. Топ-10 товаров по выручке', subtitle_style))
                elements.append(Paragraph(
                    'Горизонтальная столбчатая диаграмма отображает товары, приносящие наибольшую выручку. '
                    'Помогает определить наиболее прибыльные позиции и оптимизировать ассортимент.',
                    desc_style
                ))
                
                product_names = [item['product__name'] or '—' for item in top_products]
                # Обрезаем длинные названия для лучшей читаемости
                product_names_short = [name[:40] + '...' if len(name) > 40 else name for name in product_names]
                product_revenues = [float(item['revenue'] or 0) for item in top_products]
                
                fig, ax = plt.subplots(figsize=(10, 5.5))
                bars = ax.barh(product_names_short, product_revenues, color='#4f63d8', alpha=0.8, edgecolor='#1a2a6c', linewidth=1.5)
                ax.set_title('Топ-10 товаров по выручке', fontsize=16, fontweight='bold', pad=15)
                ax.set_xlabel('Выручка (₽)', fontsize=12, fontweight='bold')
                ax.set_ylabel('Товар', fontsize=12, fontweight='bold')
                ax.grid(True, alpha=0.3, axis='x', linestyle='--')
                
                # Добавляем значения на столбцы
                for i, (bar, revenue) in enumerate(zip(bars, product_revenues)):
                    width = bar.get_width()
                    ax.text(width, bar.get_y() + bar.get_height()/2.,
                           f'{revenue:.2f} ₽',
                           ha='left', va='center', fontweight='bold', fontsize=9)
                
                plt.tight_layout()
                
                chart4 = BytesIO()
                plt.savefig(chart4, format='png', dpi=110, bbox_inches='tight')
                plt.close()
                chart4.seek(0)
                
                elements.append(RLImage(chart4, width=9*inch, height=4.8*inch))
            
            # Таблица топ товаров
            if top_products:
                elements.append(PageBreak())
                elements.append(Paragraph('Детализация топ-товаров', subtitle_style))
                elements.append(Paragraph(
                    'Подробная таблица с данными о наиболее прибыльных товарах за выбранный период.',
                    desc_style
                ))
                elements.append(Spacer(1, 0.2*inch))
                
                table_data = [['Товар', 'Выручка', 'Количество', 'Заказов']]
                for item in top_products:
                    table_data.append([
                        item['product__name'] or '—',
                        f"{float(item['revenue'] or 0):.2f} ₽",
                        str(item['quantity'] or 0),
                        str(item['orders'] or 0)
                    ])
                
                # Улучшенная таблица с красивым форматированием
                table = Table(table_data, colWidths=[4*inch, 1.5*inch, 1.5*inch, 1.5*inch])
                table.setStyle(TableStyle([
                    # Заголовок
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a2a6c')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('FONTNAME', (0, 0), (-1, 0), title_font),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('FONTWEIGHT', (0, 0), (-1, 0), 'BOLD'),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('TOPPADDING', (0, 0), (-1, 0), 12),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
                    # Данные
                    ('ALIGN', (0, 1), (0, -1), 'LEFT'),
                    ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
                    ('FONTNAME', (0, 1), (-1, -1), normal_font),
                    ('FONTSIZE', (0, 1), (-1, -1), 10),
                    ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
                    ('TOPPADDING', (0, 1), (-1, -1), 8),
                    ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
                    # Чередующиеся цвета строк
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
                    # Границы
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
                    ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#1a2a6c')),
                ]))
                elements.append(table)
                
                # Итоговая статистика
                total_revenue = sum(float(item['revenue'] or 0) for item in top_products)
                total_quantity = sum(item['quantity'] or 0 for item in top_products)
                total_orders = sum(item['orders'] or 0 for item in top_products)
                
                elements.append(Spacer(1, 0.3*inch))
                summary_style = ParagraphStyle(
                    'Summary',
                    parent=styles['Normal'],
                    fontSize=11,
                    textColor=colors.HexColor('#1a2a6c'),
                    spaceAfter=6,
                    fontName=title_font
                )
                elements.append(Paragraph(f'<b>Итого по топ-10 товарам:</b>', summary_style))
                elements.append(Paragraph(f'Общая выручка: {total_revenue:.2f} ₽', desc_style))
                elements.append(Paragraph(f'Общее количество проданных единиц: {total_quantity}', desc_style))
                elements.append(Paragraph(f'Количество уникальных заказов: {total_orders}', desc_style))
            
            # Собираем PDF
            doc.build(elements)
            buffer.seek(0)
            
            response = HttpResponse(buffer.read(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename=dashboard_report_{timezone.now().strftime("%Y%m%d")}.pdf'
            return response
            
        except ImportError as e:
            from django.contrib import messages
            messages.error(request, _('Библиотеки для PDF не установлены. Установите: pip install reportlab matplotlib'))
            return self.get(request)
        except Exception as e:
            from django.contrib import messages
            messages.error(request, _('Ошибка при создании PDF: {}').format(str(e)))
            return self.get(request)

    def _export_all_metrics(self, request):
        """Экспорт всех метрик в Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment
            
            wb = Workbook()
            wb.remove(wb.active)
            
            styles = _get_excel_styles()
            days = int(request.POST.get('days', 30))
            start_date = timezone.now() - timedelta(days=days)
            
            # Лист 1: Продажи
            ws_sales = wb.create_sheet("Продажи")
            headers = ['Дата', 'Заказов', 'Выручка', 'Средний чек']
            _apply_excel_header_style(ws_sales, headers, styles)
            
            daily_data = (
                Order.objects.filter(
                    order_date__gte=start_date,
                    status='delivered'
                )
                .annotate(day=TruncDay('order_date'))
                .values('day')
                .annotate(
                    orders=Count('id'),
                    revenue=Sum('total_amount')
                )
                .order_by('day')
            )
            
            daily_data_list = list(daily_data)
            row = 2
            for idx, item in enumerate(daily_data_list):
                revenue = float(item['revenue'] or 0)
                orders = item['orders']
                avg = revenue / orders if orders > 0 else 0
                ws_sales.cell(row=row, column=1, value=item['day'].strftime('%d.%m.%Y'))
                ws_sales.cell(row=row, column=2, value=orders)
                ws_sales.cell(row=row, column=3, value=revenue).number_format = '#,##0.00 ₽'
                ws_sales.cell(row=row, column=4, value=avg).number_format = '#,##0.00 ₽'
                
                _apply_excel_row_style(ws_sales, row, len(headers), styles, is_alt=(idx % 2 == 0))
                ws_sales.cell(row=row, column=1).alignment = Alignment(horizontal='left')
                ws_sales.cell(row=row, column=2).alignment = Alignment(horizontal='center')
                ws_sales.cell(row=row, column=3).alignment = Alignment(horizontal='right')
                ws_sales.cell(row=row, column=4).alignment = Alignment(horizontal='right')
                
                row += 1
            
            # Итоговая строка для продаж
            if row > 2:
                total_row = row
                total_orders = sum(item['orders'] for item in daily_data_list)
                total_revenue = sum(float(item['revenue'] or 0) for item in daily_data_list)
                total_avg = total_revenue / total_orders if total_orders > 0 else 0
                
                ws_sales.cell(row=total_row, column=1, value='ИТОГО')
                ws_sales.cell(row=total_row, column=2, value=total_orders)
                ws_sales.cell(row=total_row, column=3, value=total_revenue).number_format = '#,##0.00 ₽'
                ws_sales.cell(row=total_row, column=4, value=total_avg).number_format = '#,##0.00 ₽'
                
                _apply_excel_row_style(ws_sales, total_row, len(headers), styles, is_total=True)
                ws_sales.cell(row=total_row, column=1).alignment = Alignment(horizontal='left')
                ws_sales.cell(row=total_row, column=2).alignment = Alignment(horizontal='center')
                ws_sales.cell(row=total_row, column=3).alignment = Alignment(horizontal='right')
                ws_sales.cell(row=total_row, column=4).alignment = Alignment(horizontal='right')
            
            # Лист 2: Товары
            ws_products = wb.create_sheet("Товары")
            headers = ['ID', 'Название', 'Категория', 'Цена', 'Продано шт.', 'Выручка', 'Заказов']
            _apply_excel_header_style(ws_products, headers, styles)
            
            revenue_expr = ExpressionWrapper(
                F('quantity') * F('price_per_unit'),
                output_field=DecimalField(max_digits=16, decimal_places=2)
            )
            products_data = list((
                OrderItem.objects.filter(
                    order__order_date__gte=start_date,
                    order__status='delivered'
                )
                .values('product__id', 'product__name', 'product__category__name', 'product__price')
                .annotate(
                    total_quantity=Sum('quantity'),
                    total_revenue=Sum(revenue_expr),
                    orders_count=Count('order', distinct=True)
                )
                .order_by('-total_revenue')
            ))
            
            row = 2
            for idx, item in enumerate(products_data):
                ws_products.cell(row=row, column=1, value=item['product__id'] or '—')
                ws_products.cell(row=row, column=2, value=item['product__name'] or '—')
                ws_products.cell(row=row, column=3, value=item['product__category__name'] or 'Без категории')
                ws_products.cell(row=row, column=4, value=float(item['product__price'] or 0)).number_format = '#,##0.00 ₽'
                ws_products.cell(row=row, column=5, value=item['total_quantity'] or 0)
                ws_products.cell(row=row, column=6, value=float(item['total_revenue'] or 0)).number_format = '#,##0.00 ₽'
                ws_products.cell(row=row, column=7, value=item['orders_count'] or 0)
                
                _apply_excel_row_style(ws_products, row, len(headers), styles, is_alt=(idx % 2 == 0))
                ws_products.cell(row=row, column=1).alignment = Alignment(horizontal='center')
                ws_products.cell(row=row, column=2).alignment = Alignment(horizontal='left')
                ws_products.cell(row=row, column=3).alignment = Alignment(horizontal='left')
                ws_products.cell(row=row, column=4).alignment = Alignment(horizontal='right')
                ws_products.cell(row=row, column=5).alignment = Alignment(horizontal='right')
                ws_products.cell(row=row, column=6).alignment = Alignment(horizontal='right')
                ws_products.cell(row=row, column=7).alignment = Alignment(horizontal='center')
                
                row += 1
            
            # Итоговая строка для товаров
            if row > 2:
                total_row = row
                ws_products.cell(row=total_row, column=1, value='ИТОГО')
                ws_products.cell(row=total_row, column=5, value=sum(item['total_quantity'] or 0 for item in products_data))
                ws_products.cell(row=total_row, column=6, value=sum(float(item['total_revenue'] or 0) for item in products_data)).number_format = '#,##0.00 ₽'
                ws_products.cell(row=total_row, column=7, value=sum(item['orders_count'] or 0 for item in products_data))
                
                _apply_excel_row_style(ws_products, total_row, len(headers), styles, is_total=True)
                ws_products.cell(row=total_row, column=1).alignment = Alignment(horizontal='left')
                ws_products.cell(row=total_row, column=5).alignment = Alignment(horizontal='right')
                ws_products.cell(row=total_row, column=6).alignment = Alignment(horizontal='right')
                ws_products.cell(row=total_row, column=7).alignment = Alignment(horizontal='center')
            
            # Лист 3: Клиенты
            ws_customers = wb.create_sheet("Клиенты")
            headers = ['ID', 'Имя пользователя', 'Email', 'Заказов', 'Выручка', 'Средний чек']
            _apply_excel_header_style(ws_customers, headers, styles)
            
            customers_data = list((
                User.objects.filter(
                    is_staff=False,
                    orders__order_date__gte=start_date,
                    orders__status='delivered'
                )
                .annotate(
                    orders_count=Count('orders', distinct=True),
                    total_revenue=Sum('orders__total_amount'),
                    avg_order_value=ExpressionWrapper(
                        Sum('orders__total_amount') / Count('orders', distinct=True),
                        output_field=DecimalField()
                    )
                )
                .filter(orders_count__gt=0)
                .order_by('-total_revenue')
            ))
            
            row = 2
            for idx, customer in enumerate(customers_data):
                ws_customers.cell(row=row, column=1, value=customer.id)
                ws_customers.cell(row=row, column=2, value=customer.username)
                ws_customers.cell(row=row, column=3, value=customer.email or '—')
                ws_customers.cell(row=row, column=4, value=customer.orders_count or 0)
                ws_customers.cell(row=row, column=5, value=float(customer.total_revenue or 0)).number_format = '#,##0.00 ₽'
                ws_customers.cell(row=row, column=6, value=float(customer.avg_order_value or 0)).number_format = '#,##0.00 ₽'
                
                _apply_excel_row_style(ws_customers, row, len(headers), styles, is_alt=(idx % 2 == 0))
                ws_customers.cell(row=row, column=1).alignment = Alignment(horizontal='center')
                ws_customers.cell(row=row, column=2).alignment = Alignment(horizontal='left')
                ws_customers.cell(row=row, column=3).alignment = Alignment(horizontal='left')
                ws_customers.cell(row=row, column=4).alignment = Alignment(horizontal='center')
                ws_customers.cell(row=row, column=5).alignment = Alignment(horizontal='right')
                ws_customers.cell(row=row, column=6).alignment = Alignment(horizontal='right')
                
                row += 1
            
            # Итоговая строка для клиентов
            if row > 2:
                total_row = row
                ws_customers.cell(row=total_row, column=1, value='ИТОГО')
                ws_customers.cell(row=total_row, column=4, value=sum(c.orders_count or 0 for c in customers_data))
                ws_customers.cell(row=total_row, column=5, value=sum(float(c.total_revenue or 0) for c in customers_data)).number_format = '#,##0.00 ₽'
                
                _apply_excel_row_style(ws_customers, total_row, len(headers), styles, is_total=True)
                ws_customers.cell(row=total_row, column=1).alignment = Alignment(horizontal='left')
                ws_customers.cell(row=total_row, column=4).alignment = Alignment(horizontal='center')
                ws_customers.cell(row=total_row, column=5).alignment = Alignment(horizontal='right')
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename=all_metrics_{timezone.now().strftime("%Y%m%d")}.xlsx'
            return response
            
        except ImportError:
            from django.contrib import messages
            messages.error(request, _('Библиотека openpyxl не установлена.'))
            return self.get(request)

    def _import_all_metrics(self, request):
        """Импорт метрик из Excel"""
        from django.contrib import messages
        messages.info(request, _('Импорт метрик будет реализован в будущих версиях.'))
        return self.get(request)

    def _export_styles_css(self, request):
        """Экспорт CSS стилей для Excel и PDF файлов"""
        css_content = """/* 
 * Стили для экспорта отчетов Жевжик
 * Используются в Excel и PDF файлах
 * Дата создания: {date}
 */

/* Основные цвета */
:root {{
    --primary-color: #1a2a6c;
    --primary-light: #4f63d8;
    --secondary-color: #ff6b9d;
    --success-color: #10b981;
    --warning-color: #f59e0b;
    --danger-color: #ef4444;
    --info-color: #3b82f6;
    --text-color: #1a2a6c;
    --text-light: #6c757d;
    --bg-light: #f0f4ff;
    --bg-alt: #f8f9fa;
    --border-color: #dee2e6;
}}

/* Стили для заголовков таблиц */
.table-header {{
    background: linear-gradient(135deg, var(--primary-color), var(--primary-light));
    color: #ffffff;
    font-weight: bold;
    font-size: 12px;
    text-align: center;
    padding: 12px;
    border: 1px solid var(--primary-color);
}}

/* Стили для строк данных */
.table-row {{
    background-color: #ffffff;
    border-bottom: 1px solid var(--border-color);
    padding: 8px;
}}

.table-row:nth-child(even) {{
    background-color: var(--bg-light);
}}

.table-row:hover {{
    background-color: var(--bg-alt);
}}

/* Стили для итоговых строк */
.table-total {{
    background-color: #e8f0fe;
    font-weight: bold;
    border-top: 2px solid var(--primary-color);
    padding: 10px;
}}

/* Стили для чисел */
.number-format {{
    text-align: right;
    font-family: 'Courier New', monospace;
    color: var(--text-color);
}}

.currency {{
    font-weight: 600;
    color: var(--success-color);
}}

/* Стили для PDF заголовков */
.pdf-title {{
    font-size: 24px;
    font-weight: bold;
    color: var(--primary-color);
    text-align: center;
    margin-bottom: 20px;
}}

.pdf-subtitle {{
    font-size: 16px;
    font-weight: 600;
    color: var(--primary-color);
    margin-top: 20px;
    margin-bottom: 10px;
}}

.pdf-description {{
    font-size: 10px;
    color: #666666;
    line-height: 1.6;
    margin-bottom: 15px;
}}

/* Стили для графиков */
.chart-container {{
    background: #ffffff;
    border: 1px solid var(--border-color);
    border-radius: 8px;
    padding: 15px;
    margin-bottom: 20px;
}}

.chart-title {{
    font-size: 16px;
    font-weight: bold;
    color: var(--primary-color);
    margin-bottom: 10px;
}}

/* Стили для таблиц в PDF */
.pdf-table {{
    width: 100%;
    border-collapse: collapse;
    margin: 15px 0;
}}

.pdf-table th {{
    background: var(--primary-color);
    color: #ffffff;
    font-weight: bold;
    padding: 12px;
    text-align: center;
    border: 1px solid var(--primary-color);
}}

.pdf-table td {{
    padding: 10px;
    border: 0.5px solid var(--border-color);
    text-align: left;
}}

.pdf-table tr:nth-child(even) {{
    background-color: var(--bg-light);
}}

.pdf-table tr:hover {{
    background-color: var(--bg-alt);
}}

/* Стили для Excel ячеек */
.excel-header {{
    background-color: var(--primary-color);
    color: #ffffff;
    font-weight: bold;
    font-size: 12px;
    text-align: center;
    vertical-align: middle;
    border: 1px solid #000000;
    padding: 8px;
}}

.excel-cell {{
    border: 1px solid var(--border-color);
    padding: 6px;
    font-size: 10px;
}}

.excel-cell-alt {{
    background-color: var(--bg-light);
}}

.excel-cell-total {{
    background-color: #e8f0fe;
    font-weight: bold;
    border-top: 2px solid var(--primary-color);
}}

/* Стили для форматирования чисел */
.number-currency {{
    text-align: right;
    font-family: 'Arial', sans-serif;
    color: var(--text-color);
    font-weight: 500;
}}

/* Стили для статусов */
.status-created {{ color: var(--text-light); }}
.status-confirmed {{ color: var(--info-color); }}
.status-ready {{ color: var(--warning-color); }}
.status-in-transit {{ color: var(--info-color); }}
.status-delivered {{ color: var(--success-color); }}
.status-cancelled {{ color: var(--danger-color); }}

/* Адаптивность */
@media print {{
    .table-row {{
        page-break-inside: avoid;
    }}
    
    .chart-container {{
        page-break-inside: avoid;
    }}
}}

""".format(date=timezone.now().strftime('%d.%m.%Y %H:%M'))
        
        response = HttpResponse(css_content, content_type='text/css; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename=export_styles_{timezone.now().strftime("%Y%m%d")}.css'
        return response

    def _export_batches_expiring_report(self, request):
        """Экспорт партий, истекающих в период"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from django.utils.dateparse import parse_date
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Истекающие партии"
            
            headers = ['ID партии', 'Товар', 'Категория', 'Количество', 'Дата производства', 'Срок годности', 'Дней до истечения', 'Статус']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='1a2a6c', end_color='1a2a6c', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            
            start_date = parse_date(request.POST.get('start_date'))
            end_date = parse_date(request.POST.get('end_date'))
            
            batches = ProductBatch.objects.filter(
                expiry_date__gte=start_date,
                expiry_date__lte=end_date
            ).select_related('product', 'product__category').order_by('expiry_date')
            
            row = 2
            for batch in batches:
                days_left = batch.days_until_expiry if batch.days_until_expiry is not None else 'N/A'
                status = 'Истекает скоро' if batch.days_until_expiry and batch.days_until_expiry <= 7 else 'Истекает'
                
                ws.cell(row=row, column=1, value=batch.id)
                ws.cell(row=row, column=2, value=batch.product.name if batch.product else '—')
                ws.cell(row=row, column=3, value=batch.product.category.name if batch.product and batch.product.category else '—')
                ws.cell(row=row, column=4, value=batch.quantity)
                ws.cell(row=row, column=5, value=batch.production_date.strftime('%d.%m.%Y') if batch.production_date else '—')
                ws.cell(row=row, column=6, value=batch.expiry_date.strftime('%d.%m.%Y') if batch.expiry_date else '—')
                ws.cell(row=row, column=7, value=days_left)
                ws.cell(row=row, column=8, value=status)
                row += 1
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename=batches_expiring_{timezone.now().strftime("%Y%m%d")}.xlsx'
            return response
            
        except Exception as e:
            from django.contrib import messages
            messages.error(request, _('Ошибка при экспорте: {}').format(str(e)))
            return self.get(request)

    def _export_batches_history_report(self, request):
        """Экспорт истории движений по партиям"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from django.utils.dateparse import parse_date
            
            wb = Workbook()
            ws = wb.active
            ws.title = "История партий"
            
            headers = ['Дата', 'Партия', 'Товар', 'Действие', 'Количество', 'Пользователь', 'Комментарий']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='1a2a6c', end_color='1a2a6c', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            
            start_date = parse_date(request.POST.get('start_date'))
            end_date = parse_date(request.POST.get('end_date'))
            
            logs = BatchAuditLog.objects.filter(
                created_at__date__gte=start_date,
                created_at__date__lte=end_date
            ).select_related('batch', 'batch__product', 'user').order_by('-created_at')
            
            row = 2
            for log in logs:
                ws.cell(row=row, column=1, value=log.created_at.strftime('%d.%m.%Y %H:%M'))
                ws.cell(row=row, column=2, value=log.batch.id if log.batch else '—')
                ws.cell(row=row, column=3, value=log.batch.product.name if log.batch and log.batch.product else '—')
                ws.cell(row=row, column=4, value=log.get_action_display() if hasattr(log, 'get_action_display') else log.action)
                quantity_change = (log.new_value - log.old_value) if log.old_value is not None and log.new_value is not None else 0
                ws.cell(row=row, column=5, value=quantity_change)
                ws.cell(row=row, column=6, value=log.user.username if log.user else 'Система')
                ws.cell(row=row, column=7, value=log.comment or '—')
                row += 1
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename=batches_history_{timezone.now().strftime("%Y%m%d")}.xlsx'
            return response
            
        except Exception as e:
            from django.contrib import messages
            messages.error(request, _('Ошибка при экспорте: {}').format(str(e)))
            return self.get(request)

    def _export_batches_losses_report(self, request):
        """Экспорт анализа потерь от просрочки"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from django.utils.dateparse import parse_date
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Потери от просрочки"
            
            headers = ['ID партии', 'Товар', 'Категория', 'Количество списано', 'Цена за единицу', 'Общая стоимость', 'Дата истечения', 'Дата списания']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='ef4444', end_color='ef4444', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            
            start_date = parse_date(request.POST.get('start_date'))
            end_date = parse_date(request.POST.get('end_date'))
            
            # Находим логи списания за период
            spoiled_logs = BatchAuditLog.objects.filter(
                action='spoiled',
                created_at__date__gte=start_date,
                created_at__date__lte=end_date
            ).select_related('batch', 'batch__product', 'batch__product__category')
            
            row = 2
            total_loss = Decimal('0')
            for log in spoiled_logs:
                if log.batch and log.batch.product:
                    # Используем old_value (количество до списания) или разницу
                    quantity_spoiled = log.old_value if log.old_value else (log.batch.quantity if log.batch.quantity > 0 else 0)
                    if quantity_spoiled == 0:
                        continue
                    
                    price = log.batch.product.price
                    loss = Decimal(str(quantity_spoiled)) * price
                    total_loss += loss
                    
                    ws.cell(row=row, column=1, value=log.batch.id)
                    ws.cell(row=row, column=2, value=log.batch.product.name)
                    ws.cell(row=row, column=3, value=log.batch.product.category.name if log.batch.product.category else '—')
                    ws.cell(row=row, column=4, value=quantity_spoiled)
                    ws.cell(row=row, column=5, value=float(price))
                    ws.cell(row=row, column=6, value=float(loss))
                    ws.cell(row=row, column=7, value=log.batch.expiry_date.strftime('%d.%m.%Y') if log.batch.expiry_date else '—')
                    ws.cell(row=row, column=8, value=log.created_at.strftime('%d.%m.%Y %H:%M'))
                    row += 1
            
            # Итоговая строка
            if row > 2:
                ws.cell(row=row, column=1, value='ИТОГО').font = Font(bold=True)
                ws.cell(row=row, column=6, value=float(total_loss)).font = Font(bold=True)
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename=batches_losses_{timezone.now().strftime("%Y%m%d")}.xlsx'
            return response
            
        except Exception as e:
            from django.contrib import messages
            messages.error(request, _('Ошибка при экспорте: {}').format(str(e)))
            return self.get(request)

    def _export_batches_by_category_report(self, request):
        """Экспорт остатков по категориям"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Остатки по категориям"
            
            headers = ['Категория', 'Товаров', 'Партий', 'Общее количество', 'Средняя цена', 'Общая стоимость']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='1a2a6c', end_color='1a2a6c', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            
            # Агрегируем данные по категориям
            from django.db.models import Avg
            categories_data = (
                ProductBatch.objects.filter(remaining_quantity__gt=0)
                .values('product__category__name')
                .annotate(
                    products_count=Count('product', distinct=True),
                    batches_count=Count('id'),
                    total_quantity=Sum('remaining_quantity'),
                    avg_price=Avg('product__price'),
                    total_value=Sum(ExpressionWrapper(F('remaining_quantity') * F('product__price'), output_field=DecimalField()))
                )
                .order_by('-total_value')
            )
            
            row = 2
            for item in categories_data:
                ws.cell(row=row, column=1, value=item['product__category__name'] or 'Без категории')
                ws.cell(row=row, column=2, value=item['products_count'] or 0)
                ws.cell(row=row, column=3, value=item['batches_count'] or 0)
                ws.cell(row=row, column=4, value=item['total_quantity'] or 0)
                ws.cell(row=row, column=5, value=float(item['avg_price'] or 0))
                ws.cell(row=row, column=6, value=float(item['total_value'] or 0))
                row += 1
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename=batches_by_category_{timezone.now().strftime("%Y%m%d")}.xlsx'
            return response
            
        except Exception as e:
            from django.contrib import messages
            messages.error(request, _('Ошибка при экспорте: {}').format(str(e)))
            return self.get(request)

